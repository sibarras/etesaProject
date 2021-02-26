from typing import Tuple
from openpyxl.reader.excel import ExcelReader
from openpyxl.styles.cell_style import CellStyle
from openpyxl.worksheet.worksheet import Worksheet
from selenium.webdriver import Chrome
from selenium.webdriver.common.by import By
from openpyxl import load_workbook

from time import sleep
from datetime import datetime
from pathlib import Path
import pandas as pd
import tkinter as tk
from tkinter import messagebox

from secrets import SECRET_PASSWORD, SECRET_USER

from data import places_id, work_tools
from works_example import works, PMAWork
from works_struc import PMAWork


class WorkFile:
    def __init__(self, name:str) -> None:
        self.name = name
        self.lines = []
        self.count = 0
        self.works_path = Path(__file__).parent
        self.file_path = self.works_path / name
        if self.file_path.exists():
            with open(self.file_path.__str__(), 'rt') as f:
                print('File opened.')
                for row in f.readlines():
                    self.lines.append(row)
                    if row == '\n':
                        self.count += 1
        else:
            self.__create_file()
    
    def refresh_lines(self):
        final_num = self.count
        num = 0
        with open(self.file_path.__str__(), 'rt') as f:
            for row in f.readlines():
                self.lines.append(row)
                if 'works_' in row :
                    num = int(row.lstrip('works_').split()[0])
                    if num > final_num:
                        final_num = num

        self.count = final_num
    
    def __create_file(self):
        f = open(self.file_path.__str__(), 'wt')
        f.close()
        print('Document created')

    def add(self, work:PMAWork):
        # Mejorar el uso de esta funcion, se llama por otra clase sin avisar
        work=work.__dict__
        with open(self.name, 'at') as f:
            w = f'works_{self.count}'
            w += " = {\n"
            for key, val in work.items():
                if f'\t"{key}"="{val}"' not in self.lines:
                    w += f'\t"{key}":"{val}",\n'
            w += "}\n\n"
            f.write(w)
            self.refresh_lines()
        self.count += 1


class OTCreator:
    def __init__(self, py_file: WorkFile, excel_file:Path) -> None:
        self.file = py_file
        self.xl_file = excel_file
        self.ot_folder:Path
        self.works = []

        wb = load_workbook(excel_file)
        self.ws:Worksheet = wb[(wb.sheetnames[0])]
        self.excel_table:Tuple[Tuple, ...] = self.ws['A15:P304']
        self.excel_index = lambda id: id+16
    
    def set_ot_in_excel(self, ot_number, id):
        assert -1<=id<289 and type(id) is int
        wb = load_workbook(self.xl_file.__str__())
        ws:Worksheet = wb[(wb.sheetnames[0])]
        for row in ws['O16:O304']:
            for cell in row:
                if cell.row == self.excel_index(id):
                    cell.value = ot_number
                    wb.save(self.xl_file)
        wb.close()
        
    
    def popup(self, title, message) -> bool:
        root = tk.Tk()
        root.withdraw()
        ans = messagebox.askyesno(title = title, message = message)
        return ans
    
    def do(self, id:str, action:str, message:str='', write_validation=True, count=0) -> str:
        sleep(0.5)
        if count <= 3:
            try:
                element = self.driver.find_element(By.XPATH, f'//*[@id="{id}"]')
                if action == 'click':
                    element.click()
                elif action == 'write':
                    element.clear()
                    element.send_keys(message)
                    if write_validation:
                        assert element.get_attribute('value') == message or element.text == message
                elif action == 'read':
                    return element.get_attribute('value')
            except KeyboardInterrupt:
                count = 4
            except Exception:
                count += 1
                sleep(count)
                self.do(id, action, message, write_validation, count)
        else:
            prompt = f"""
                Existe un problema.
                No se pudo interactuar con el elemento.
                Presione No para terminar o Si en caso de ya haberlo hecho de forma manual.
                Elemento {id} con el objetivo {action}.
            """
            ans = self.popup(title='Error de interaccion.', message=prompt)

            if ans:
                return '' # Retorna un string porque en la opcion read hace eso.
            else:
                self.driver.quit()

    def create_ots(self, works: list):
        with Chrome() as driver:
            self.driver = driver
            self.driver.implicitly_wait(60)
            # Navegar a la pagina de maximo
            self.driver.get('http://pegasuswebprod.etesa.com.pa/maximo/webclient/login/login.jsp')
            
            self.login()

            for work in works:
                if works.index(work) == 0:
                    self.first_ot()
                else:
                    self.next_ot()
                
                try:
                    self.ot_page_creator(work)
                    self.planning_page_creator(work)
                except Exception as e:
                    print(f'NO se puedo crear la ot con id {work.id}. Error:', e)

                validation = self.review()
                if validation:
                    self.save_and_send() 
                else:
                    self.signout()
                    self.do('m96ad0396-pb', 'click')
                    sleep(2)
                    driver.quit()
                    return


            self.signout()
            
        
    def login(self):
        print('Iniciando Sesion')
        self.do('username', 'write', SECRET_USER)
        self.do('password', 'write', SECRET_PASSWORD)
        self.do('loginbutton', 'click')

    def first_ot(self):
        print('Creando OT')
        self.do('QuickInsert_WOTRACK', 'click')

    def next_ot(self):
        print('Creando OT')
        self.do('toolactions_INSERT-tbb_image', 'click')

    def ot_page_creator(self, work:PMAWork):
        print('Creando Primera Pagina')
        sleep(1)
        self.do('mad3161b5-tb2', 'write', work.titulo)
        # Obtener el numero de OT
        sleep(2)
        ot_number = self.do('mad3161b5-tb', 'read')
        assert type(int(ot_number)) is int
        work.ot = ot_number
        print(f'Titulo: {work.titulo}\nFecha:{work.inicio}\nOT:{work.ot}\n\n')
        self.works.append(work)
        try:
            print('Guardando en txt')
            self.save_work_txt(work)
        except Exception as e:
            print('No se pudo guardar el .txt.', e)
        try:
            print('Guardando en .py')
            self.file.add(work)
        except Exception as e:
            print('No se pudo guardar el .py.', e)
        try:
            print('Guardando en excel')
            self.set_ot_in_excel(work.ot, work.id)
        except Exception as e:
            print('No se pudo guardar el excel.', e)

        # Cargar datos de trabajo
        self.do('mad3161b5-tb2', 'write', work.titulo)

        # Ubicacion del Equipo
        self.do('m7b0033b9-img', 'click')
        self.do('LOCATIONS_locations0', 'click')
        self.do('lookup_page1_tfrow_[C:0]_txt-tb', 'write', places_id[self.__normalize(work.ubicacion)])
        self.do('lookup_page1-ti2_img', 'click')
        self.do('lookup_page1_tdrow_[C:0]_ttxt-lb[R:0]', 'click')

        # Cargamos datos de cuenta y el resto
        self.do("ma26371c5-tb", 'write', work.cuenta)
        self.do("me2096203-tb", 'write', work.tipo)
        sleep(1)
        self.do("mc8f7970f-tb", 'write', work.clasificacion)
        sleep(1)
        self.do("m526c8119-tb", 'write', work.libranza)
        self.do("m950e5295-tb", 'write', work.prioridad)
        self.do("m651c06b0-tb", 'write', work.inicio)
        sleep(1)
        self.do("m8c7fa385-tb", 'write', work.duracion)
        self.do("mb2eb834-tb", 'write', work.supervisor)
        sleep(1)

    def planning_page_creator(self, work:PMAWork):
        print('Creando segunda pagina')

        # Accedemos a planes
        self.do("m356798d1-tab", 'click')

        # Generar tarea
        self.do('mbb442a0c_bg_button_addrow-pb', 'click')
        self.do("mb804724a-tb2", 'write', work.titulo)
        self.do("m558ec708-tb", 'write', work.inicio)
        self.do("mc86e86c5-tb", 'write', work.duracion)

        # Cargar Mano de Obra
        self.do("m95d763ff-tab", 'click')
        for worker_id in work.mano_de_obra:
            self.do("m5e4b62f0_bg_button_addrow-pb", 'click')
            self.do("m7171111e-tb", 'write', worker_id, write_validation=False)

        # Cargar Herramientas
        self.do("m716a3889-tab", 'click')
        for tool_id in work_tools[work.herramientas]:
            self.do("maa88f684_bg_button_addrow-pb", 'click')
            self.do("m8b93c9ab-tb", 'write', tool_id, write_validation=False)

    def review(self):
        return self.popup(
            title = 'Revisión del documento',
            message = 'Revisa el documento detenidamente. Presiona si para continuar, o presiona no para descartar.',
        )

    def save_and_send(self):
        print('Guardando y Enrutando')
        self.do('toolactions_SAVE-tbb_image', 'click')
        self.do('ROUTEWF_WOETESA_-tbb_image', 'click')
    
    def save_work_txt(self, work: PMAWork):
        timestamp = datetime.timestamp(datetime.now())
        with open(self.ot_folder/f'work_{work.id}_{timestamp}.txt', 'wt') as f:
            text = ''
            for key, val in work.__dict__.items():
                text += f'{key}: {val}\n'
            f.write(text)

    def signout(self):
        print('Cerrando Sesion')
        self.do("titlebar_hyperlink_8-lbsignout", 'click')
        
    def add_works_to_file(self, file: WorkFile, work: PMAWork) -> None:
        print('Añadiendo archivos a base de datos')
        if len(self.works):
            file.add(work)
    
    def get_works_dataframe(self) -> pd.DataFrame:
        self.df = pd.DataFrame([work.__dict__ for work in self.works])
        return self.df

    def __normalize(self, sub):
        return sub.title()\
                .replace('Iii', 'III')\
                .replace('Ii', 'II')\
                .replace('De', 'de')\
                .replace('Sanchez', 'Sánchez')\
                .strip()


if __name__ == '__main__':
    python_file = Path(__file__).parent/'results'/'works.py'
    excel_file = Path(__file__).parent/'excel'/'PMA2021_Pruebas_Y_Mediciones_VF.xlsx'
    # prueba
    file = WorkFile(python_file)
    maximo = OTCreator(file, excel_file)
    # maximo.create_ots(works)
    # print(maximo.get_works_dataframe())