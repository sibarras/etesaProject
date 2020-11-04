import os

if 'accounts.db' not in os.listdir('./database'):
    readExcel()

def readExcel():

    from openpyxl import load_workbook
    import sqlite3
    import pandas as pd

    PATH = './Excel Books/'
    #abro el segundo archivo y lo declaro como accounts workbook
    accounts_filename = 'accounts_PyM.xlsx'
    accounts_wb = load_workbook(PATH + accounts_filename, read_only=True, data_only=True)

    # Abro la hoja de Cuentas de ETESA
    accounts_ws = accounts_wb['CUENTAS ETESA']

    # Para observar las hojas dentro de un workbook
    #print(accounts_wb.sheetnames)


    # Creo mi encabezado para mi lista de cuentas
    account_list = []
    account_list.append([])
    for col in accounts_ws['4'][1:]:
        account_list[-1].append(col.value)


    # Filtrado por el valor de la columna Tipo
    coord_index = account_list[0].index('TIPO') + 1
    cost_index = account_list[0].index('CENTRO DE COSTO') + 1
    for row in accounts_ws.rows:
        if row[coord_index].value == 'MED' and 'PM-MED-' in row[cost_index].value:
            account_list.append([])
            for box in row[1:]:
                account_list[-1].append(box.value)


    # diccionario para la data de las ubicaciones
    place_data = {}
    place_listindex = account_list[0].index('UBICAC')
    for place in account_list:
        place_data[place[place_listindex]] = place

    # Creo un dataframe con los valores utiles del programa
    df = pd.DataFrame(data=account_list[1:], columns=account_list[0])

    # Guardo este dataframe en una base de datos.
    try:
        conn = sqlite3.connect('./database/accounts.db')
        df.to_sql('accounts', conn)
    except:
        print('[ERROR]: La base de datos de cuentas ya existe.')
    finally:
        conn.close()
