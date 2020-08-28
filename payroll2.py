from openpyxl import load_workbook
import calendar
import datetime
PATH = './Excel Books/'

#abro el segundo archivo y lo declaro como accounts workbook
accounts_filename = 'accounts_PyM.xlsx'
accounts_wb = load_workbook(PATH + accounts_filename, data_only=True)

# Abro la hoja de Cuentas de ETESA
accounts_ws = accounts_wb['CUENTAS ETESA']

# Para observar las hojas dentro de un workbook
#accounts_wb.get_sheet_names()

# Otra forma de abrir las hojas
# accounts_ws2 = accounts_wb.get_sheet_by_name('CUENTAS ETESA')


# Creo mi encabezado para mi lista de cuentas
account_list = []
account_list.append([])
for col in accounts_ws['4:4'][1:]:
    account_list[-1].append(col.value)


# Filtrado por el valor de la columna Tipo
coord_index = account_list[0].index('TIPO') + 1
cost_index = account_list[0].index('CENTRO DE COSTO') + 1
for row in accounts_ws.rows:
    if row[coord_index].value == 'MED' and 'PM-MED-' in str(row[cost_index].value):
        account_list.append([])
        for box in row[1:]:
            account_list[-1].append(box.value)

# Para ver los datos de la lista
# for row in account_list:
#     print()
#     for col in row:
#         print(col, end='\t')

# diccionario para la data de las ubicaciones
place_data = {}
place_listindex = account_list[0].index('UBICAC')
for place in account_list:
    place_data[place[place_listindex]] = place
print(place_data.keys())
# Otros datos a utilizar
tests = [
    'Tiempos de Operacion',         #0
    'Resistencia de Contacto',      #1
    'Resistencia de Aislamiento',   #2
    'Resistencia de Devanados',     #3
    'Relacion de Vueltas',          #4
    'Factor de Potencia',           #5
    'Analisis de Gas SF6',          #6
    'Analis de Calidad de Energia', #7
    'Corriente de Excitacion',      #8
    'Saturacion',                    #9
    'Alarmas',                      #10
    'Disparos',                     #11
    'Bloqueos',                     #12
    'Tratamiento de gas SF6',       #13 
]
equipmentTests = {
    'PT': [tests[i] for i in [5, 2]],
    'CT': [tests[i] for i in [2, 3, 4, 5, 8, 9]],
    'INT': [tests[i] for i in [0, 1, 6, 10, 12]],
    'TX': [tests[i] for i in [5, 8, 4, 3, 10, 11]],
    'RX': [tests[i] for i in [5, 10, 11]]
}
se = 'Subestación '
nombre_SE = {
    'SPANAMA2': se+'Panamá II',
    'SPANAMA': se+'Panamá',
    'SCHORRE': se+'Chorrera',
    'SCACERE': se+'Cáceres',
    'SLLSANC': se+'Llano Sánchez',
    'SSTA RITA': se+'Santa Rita',
    'SVELADER': se+'Veladero',
    'SMDN': se+'Mata de Nance',
    'SPROGRE': se+'Progreso',
    'SGUASQU': se+'Guasquitas',
    'SCALDER': se+'Caldera',
    'SCHAZUL': se+'Charco Azul',
    'SCHANGU': se+'Changuinola',
    'SCHILIB': se+'Chilibre',
    'CVALES': se+'Los Valles',
    'CESTREL': se+'La Etrella',
    'CPACORA': se+'Pacora',
    'CBLM1': se+'Bahía Las Minas',
    'CPAN AM': se+'Panam',
    'CESTI': se+'Estí',
    'CFORT': se+'Fortuna',
    'CBLM': se+'Bahía Las Minas',
    'CBAYAN': se+'Bayano',
    'SBARTOLO': se+'San Bartolo',
    'SHIGO': se+'El Higo'
}
if list(place_data.keys())[1:] == list(nombre_SE.keys()):
    print('diccionario correcto')

spanishDict = {
    'Monday': 'Lunes',
    'Tuesday': 'Martes',
    'Wednesday': 'Miercoles',
    'Thursday': 'Jueves',
    'Friday': 'Viernes',
    'Saturday': 'Sabado',
    'Sunday': 'Domingo',
    'January': 'Enero',
    'February': 'Febrero',
    'March': 'Marzo',
    'April': 'Abril',
    'May': 'Mayo',
    'June': 'Junio',
    'July': 'Julio',
    'August': 'Agosto',
    'September': 'Septiembre',
    'October': 'Octubre',
    'November': 'Noviembre',
    'December': 'Diciembre'
}

# Abro el primer archivo y lo declaro como payroll workbook
payroll_filename = 'payroll_layout.xlsx'
payroll_wb = load_workbook(PATH + payroll_filename)

# Para la distribucion del tiempo
distribution_ws = payroll_wb['DISTRIBUCION']

# Para las horas extras
extra_ws = payroll_wb['EXTRAS']

# Vamos a escribir primeramente las horas extras * traducir
nombre_cell = 'I11'
cedula_cells = [f'{chr(i)}9' for i in range(ord('T'), ord('Z')+1)]\
                + [f'A{chr(i)}9' for i in range(ord('A'), ord('D')+1)]
numEmpleado_cell = 'X11'
gerencia_cell = 'I12'
depto_cell = 'Z12'
cargo_cell = 'I13'
periodo_cells = ['X13','AB13']

lugar_rows = [str(col) for col in range(17, 21+1)]
dias_cols = [f'{chr(i)}' for i in range(ord('R'), ord('Z')+1)]\
            + [f'A{chr(i)}' for i in range(ord('A'), ord('G')+1)]
codigoLugar_col = 'C'

tarea_rows = [str(col) for col in range(25, 33+1)]
tarea_col = 'C'
rangoHoras_cols = ['AB','AF']

enfermedad_row = '37'
fiestaDuelo_row = '38'
accidente_row = '39'
adiestram_row = '40'
vacaciones_row = '41'
compensatiorio_row = '42'
sindical_row = '43'
otros_row = '44'

# Tienes que anadir aqui los cambios a hacer en distribucion sheet
def dist_data():
    nombre_cell = 'F:7'
    cedula_cells = [f'{chr(i)}:6' for i in range(ord('S'), ord('Z')+1)]\
                    + [f'A{chr(i)}:6' for i in range(ord('A'), ord('C')+1)]
    numEmpleado_cell = 'W:7'
    gerencia_cell = 'F:8'
    depto_cell = 'W:8'
    cargo_cell = 'F:9'
    periodo_cells = ['Y:9','AB:9']

    lugarInit_cols = [col for col in range(17, 21+1)]
    dias_cols = [f'{chr(i)}' for i in range(ord('Q'), ord('Z')+1)]\
                + [f'A{chr(i)}' for i in range(ord('A'), ord('H')+1)]
    codigoLugar_col = ''

    tareaInit_rows = [col for col in range(25, 33+1)]
    tarea_col = ''
    inicioFin_cols = ['','']

    enfermedad_row = '37'
    fiestaDuelo_row = '38'
    accidente_row = '39'
    adiestram_row = '40'
    vacaciones_row = '41'
    compensatiorio_row = '42'
    sindical_row = '43'
    otros_row = '44'


# datos del usuario * Utilizar otro archivo con esta data
nombre = 'Samuel Eliezer Ibarra Solis'
cedula = '8-892-2460'
num_empleado = '27720'
gerencia = 'Proteccion y Comunicacion'
depto = 'Pruebas y Mediciones'
cargo = 'Especialista en Pruebas y Mediciones'

# Ahora introducir estos datos en el extra worksheet
extra_ws[nombre_cell] = nombre.upper()
extra_ws[numEmpleado_cell] = num_empleado
extra_ws[gerencia_cell] = gerencia.upper()
extra_ws[depto_cell] = depto.upper()
extra_ws[cargo_cell] = cargo.upper()

# cedula
ced_parts = cedula.split('-')
ced_slots = ([1,5,5] if len(ced_parts)== 3 else [1,2,3,5])
for cell in cedula_cells:
    extra_ws[cell] = 0

# aplicar
for part in reversed(ced_parts):
    for num in reversed(part):
        extra_ws[cedula_cells[-1]] = num
        cedula_cells.pop()
    while len(cedula_cells) > sum(ced_slots[:-1]):
        cedula_cells.pop()
    ced_slots.pop()

# Datos periodo
quincena, mes = 2, 8

# Aplicar
year = int(str(datetime.datetime.now()).split()[0].split('-')[0])
nombre_mes = calendar.month_name[mes]
dia_fin_mes = calendar.monthlen(year, mes)
nombre_dia = lambda d: calendar.day_name[calendar.weekday(year, mes, d)]
periodo_planilla = lambda : 1 if quincena == 1 else 16,\
                    lambda : f'15 de {nombre_mes}' if quincena == 1 else f'{dia_fin_mes} de {nombre_mes}' 

# aplico datos en worksheet
extra_ws[periodo_cells[0]] = periodo_planilla[0]()
extra_ws[periodo_cells[1]] = periodo_planilla[1]().upper()

# Datos de trabajo realizado
# Debes hacerle un sort a cada parte de los trabajos por los dias de referencia
works = {
    'place':['SPANAMA', 'SPANAMA2', 'SPANAMA'],
    'day':[23, 26, 30],
    'init':['3:30', '6:00', '3:30'],
    'end':['6:45', '7:00', '4:30'],
    'equip':['CT', 'TX', 'PT']
}

# comprobar datos de trabajo
numWorks = 0
for data in works.values():
    if numWorks == 0: numWorks = len(data)
    if len(data) != numWorks:
        print('ERROR EN CANTIDAD DE DATOS EN LOS TRABAJOS')
        exit()

for day in works['day']:
    if 1 > day or day > 15 and quincena == 1 or dia_fin_mes < day or day < 15 and quincena == 2:
        print('ERROR EN DIA ESTIPULADO POR ESTAR FUERA DEL PERIODO')
        exit()

# codigo de lugar y dias de trabajo
codes_in_ws = []
horas_trabajo = lambda init, end : (int(end.split(':')[0]) + int(end.split(':')[1])/60)\
                                    - (int(init.split(':')[0]) + int(init.split(':')[1])/60)

for n in range(numWorks):
    # Codigos
    if place_data[works['place'][n]][-1] not in codes_in_ws:
        extra_ws[codigoLugar_col+lugar_rows[n]] = place_data[works['place'][n]][-1]  # tomado del otro ws
        codes_in_ws.append(place_data[works['place'][n]][-1])
    # dias
    code_index = codes_in_ws.index(place_data[works['place'][n]][-1])
    if quincena == 1:
        extra_ws[dias_cols[works['day'][n]-1]+lugar_rows[code_index]]\
             = horas_trabajo(works['init'][n], works['end'][n])
    elif quincena == 2:
        extra_ws[dias_cols[works['day'][n]-16]+lugar_rows[code_index]]\
             = horas_trabajo(works['init'][n], works['end'][n])

# justificacion y horas de justificacion
justificacion_char = []
for n in range(numWorks):
    pruebas = ''
    testList = equipmentTests[works['equip'][n]]
    for test in testList:
        if test == testList[-1]:
            pruebas = pruebas[:-2] + ' y ' + test
            continue
        pruebas += test + ', '
    justificacion_char.append(f'{nombre_dia(works["day"][n])} {works["day"][n]}\
                               de {nombre_mes} de {year}: Pruebas de {pruebas} \
                               en {nombre_SE[works["place"][n]]}')

# aplicar junto con las horas
for n in range(numWorks):
    extra_ws[tarea_col+tarea_rows[n]] = justificacion_char[n]
    extra_ws[rangoHoras_cols[0]+tarea_rows[n]] = works['init'][n]
    extra_ws[rangoHoras_cols[1]+tarea_rows[n]] = works['end'][n]

# horas no laboradas
print('horas no laboradas en desarrollo')
# aplicar

# guardar en nuevo documento
payroll_wb.save(PATH+'tryDocument.xlsx')


